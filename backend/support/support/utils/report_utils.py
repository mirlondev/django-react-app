# utils/report_utils.py
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from django.utils import timezone
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from datetime import datetime
import pandas as pd
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import calendar






def export_monthly_report_excel(interventions, month, year):
    """
    Generate a professional Excel monthly report with detailed intervention data
    using FCFA currency and French date formats
    """
    # Create a workbook and get the active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rapport {month}-{year}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=14, color="366092")
    subtitle_font = Font(bold=True, italic=True, size=10)
    
    # Use FCFA currency format
    currency_format = '#,##0.00 "FCFA"'
    
    # French date and time formats
    date_format = "DD/MM/YYYY"
    time_format = "HH:MM"
    
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Add title and metadata in French
    month_name = [
        "Janvier", "Février", "Mars", "Avril", "Mai", "Juin",
        "Juillet", "Août", "Septembre", "Octobre", "Novembre", "Décembre"
    ][month-1]
    
    ws.merge_cells('A1:L1')
    ws['A1'] = f"RAPPORT MENSUEL D'INTERVENTIONS - {month_name.upper()} {year}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:L2')
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    ws['A2'].font = subtitle_font
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Prepare data for the table
    data = []
    for intervention in interventions:
        # Get related objects safely
        ticket = getattr(intervention, 'ticket', None)
        client = getattr(ticket, 'client', None) if ticket else None
        technician = getattr(intervention, 'technician', None)
        tech_user = getattr(technician, 'user', None) if technician else None
        
        # Format times in French format
        start_time = intervention.start_time.strftime('%H:%M') if intervention.start_time else "N/A"
        end_time = intervention.end_time.strftime('%H:%M') if intervention.end_time else "N/A"
        
        # Calculate total time if not already available
        total_time = getattr(intervention, 'calculate_total_time', None)
        if total_time and callable(total_time):
            total_time = total_time()
        else:
            total_time = (intervention.hours_worked or 0) + (intervention.travel_time or 0)
        
        # Get client and technician names
        client_name = "N/A"
        if client:
            client_name = getattr(client, 'company', None) or f"{getattr(getattr(client, 'user', None), 'first_name', '')} {getattr(getattr(client, 'user', None), 'last_name', '')}".strip()
        
        tech_name = "N/A"
        if tech_user:
            tech_name = f"{getattr(tech_user, 'first_name', '')} {getattr(tech_user, 'last_name', '')}".strip()
        
        # Add row data with French date format
        data.append({
            'Date': intervention.intervention_date.strftime('%d/%m/%Y') if intervention.intervention_date else "N/A",
            'Client': client_name,
            'Technicien': tech_name,
            'ID Ticket': f"TKT-{ticket.id}" if ticket else "N/A",
            'Heure début': start_time,
            'Heure fin': end_time,
            'Heures travaillées': intervention.hours_worked or 0,
            'Temps déplacement': intervention.travel_time or 0,
            'Temps total': total_time,
            'Frais transport': intervention.transport_cost or 0,
            'Frais supplémentaires': intervention.additional_costs or 0,
            'Coût total': intervention.total_cost or 0,
            'Statut': getattr(intervention, 'status', 'N/A').capitalize(),
            'Type service': getattr(ticket, 'service_type', 'N/A') if ticket else 'N/A',
            'Priorité': getattr(ticket, 'priority', 'N/A').capitalize() if ticket else 'N/A',
            'Localisation': getattr(client, 'address', 'N/A') if client else 'N/A',
            'Contact': getattr(client, 'phone', 'N/A') if client else 'N/A',
        })
    
    # Create DataFrame
    df = pd.DataFrame(data)
    
    # Define column order and headers in French
    column_order = [
        'Date', 'Client', 'Technicien', 'ID Ticket', 'Type service', 'Priorité',
        'Heure début', 'Heure fin', 'Heures travaillées', 'Temps déplacement', 'Temps total',
        'Frais transport', 'Frais supplémentaires', 'Coût total', 'Statut', 'Localisation', 'Contact'
    ]
    
    # Reorder columns
    df = df[column_order]
    
    # Write headers
    for col_idx, column_name in enumerate(column_order, 1):
        cell = ws.cell(row=4, column=col_idx, value=column_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data rows
    for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 5):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Apply number formatting for currency and time columns
            if column_order[col_idx-1] in ['Frais transport', 'Frais supplémentaires', 'Coût total']:
                cell.number_format = currency_format
            elif column_order[col_idx-1] in ['Heures travaillées', 'Temps déplacement', 'Temps total']:
                cell.number_format = '0.00" heures"'
    
    # Adjust column widths
    column_widths = {
        'Date': 12,
        'Client': 25,
        'Technicien': 20,
        'ID Ticket': 12,
        'Type service': 15,
        'Priorité': 10,
        'Heure début': 10,
        'Heure fin': 10,
        'Heures travaillées': 12,
        'Temps déplacement': 12,
        'Temps total': 12,
        'Frais transport': 15,
        'Frais supplémentaires': 15,
        'Coût total': 15,
        'Statut': 12,
        'Localisation': 30,
        'Contact': 15
    }
    
    for col_idx, column_name in enumerate(column_order, 1):
        column_letter = ws.cell(row=4, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = column_widths[column_name]
    
    # Add summary section in French
    summary_row = len(data) + 6
    
    # Add summary headers
    ws.cell(row=summary_row, column=1, value="RÉSUMÉ").font = title_font
    summary_row += 1
    
    # Add summary data
    summary_data = [
        ["Total Interventions", len(interventions)],
        ["Heures totales", df['Temps total'].sum()],
        ["Total Frais Transport", df['Frais transport'].sum()],
        ["Total Frais Supplémentaires", df['Frais supplémentaires'].sum()],
        ["Revenu Total", df['Coût total'].sum()],
        ["Temps moyen par intervention", df['Temps total'].mean() if len(interventions) > 0 else 0],
        ["Revenu moyen par intervention", df['Coût total'].mean() if len(interventions) > 0 else 0],
    ]
    
    for i, (label, value) in enumerate(summary_data, summary_row):
        ws.cell(row=i, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=i, column=2, value=value)
        
        if "Frais" in label or "Revenu" in label or "Coût" in label:
            cell.number_format = currency_format
        elif "Heures" in label or "Temps" in label:
            cell.number_format = '0.00" heures"'
    
    # Add status breakdown in French
    status_row = summary_row + len(summary_data) + 2
    ws.cell(row=status_row, column=1, value="RÉPARTITION PAR STATUT").font = title_font
    status_row += 1
    
    status_counts = df['Statut'].value_counts()
    for i, (status, count) in enumerate(status_counts.items(), status_row):
        ws.cell(row=i, column=1, value=status)
        ws.cell(row=i, column=2, value=count)
    
    # Freeze panes for easy navigation
    ws.freeze_panes = "A5"
    
    # Save to BytesIO buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return buffer

def export_intervention_pdf(intervention):
    """Generate a professional PDF report for an intervention with FCFA currency and French format"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36
    )

    elements = []
    styles = getSampleStyleSheet()
    
    # Styles
    title_style = ParagraphStyle(
        'Title', parent=styles['Heading1'], fontSize=16,
        alignment=1, textColor=colors.HexColor('#2c3e50'), spaceAfter=12
    )
    section_style = ParagraphStyle(
        'Section', parent=styles['Heading2'], fontSize=10,
        textColor=colors.HexColor('#2c3e50'), spaceAfter=4
    )
    normal_style = ParagraphStyle(
        'Normal', parent=styles['Normal'], fontSize=8, spaceAfter=2
    )

    # --- Header with Logo & Title ---
    logo_path = "logo.png"  # Mettre le chemin réel de ton logo
    try:
        logo = Image(logo_path, width=1*inch, height=1*inch)
    except:
        logo = Paragraph(" ", normal_style)  # Si pas de logo

    title = Paragraph("RAPPORT D'INTERVENTION", title_style)
    header_table = Table([[logo, title]], colWidths=[1*inch, 5*inch])
    header_table.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('ALIGN',(1,0),(1,0),'CENTER'),
        ('LEFTPADDING',(0,0),(-1,-1),0),
        ('RIGHTPADDING',(0,0),(-1,-1),0)
    ]))
    elements.append(header_table)
    elements.append(Spacer(1,12))

    # --- Company & Client Info ---
    client = getattr(intervention.ticket,'client',None)
    ticket_id = getattr(intervention.ticket,'id','N/A')
    username = getattr(getattr(client,'user',None),'username','N/A')
    address = getattr(client,'address','N/A') if client else 'N/A'
    contact = getattr(client,'phone','N/A') if client else 'N/A'

    company_info = [
        Paragraph("<b>Prestataire de Services:</b>", normal_style),
        Paragraph("Nom de Votre Société", normal_style),
        Paragraph("123 Rue de l'Entreprise, Ville, Pays", normal_style),
        Paragraph("+221 33 123 45 67 | contact@votresociete.com", normal_style),
        Paragraph("ID Fiscal: M2005110000307071 | Reg: CG/PNR/10 B 1445", normal_style),
    ]
    client_info = [
        Paragraph("<b>Informations Client:</b>", normal_style),
        Paragraph(username, normal_style),
        Paragraph(address, normal_style),
        Paragraph(contact, normal_style),
        Paragraph(f"Référence Ticket: #{ticket_id}", normal_style)
    ]

    info_table = Table([[company_info, client_info]], colWidths=[3*inch, 3*inch])
    info_table.setStyle(TableStyle([
        ('VALIGN',(0,0),(-1,-1),'TOP'),
        ('LEFTPADDING',(0,0),(-1,-1),4),
        ('RIGHTPADDING',(0,0),(-1,-1),4)
    ]))
    elements.append(info_table)
    elements.append(Spacer(1,12))

    # --- Intervention Details Table ---
    technician = getattr(intervention, 'technician', None)
    tech_name = ""
    if technician:
        tech_user = getattr(technician,'user',None)
        if tech_user:
            tech_name = f"{getattr(tech_user,'first_name','')} {getattr(tech_user,'last_name','')}".strip()

    # Use French date format
    date_str = getattr(intervention,'intervention_date',datetime.now()).strftime('%d/%m/%Y')
    start_str = getattr(intervention,'start_time','N/A')
    if start_str != 'N/A':
        start_str = start_str.strftime('%H:%M')
    end_str = getattr(intervention,'end_time','N/A')
    if end_str != 'N/A':
        end_str = end_str.strftime('%H:%M')
    hours = getattr(intervention,'hours_worked','0')
    travel = getattr(intervention,'travel_time','0')
    total_cost = getattr(intervention,'total_cost','0')

    # Status mapping to French
    status_mapping = {
        'pending': 'En attente',
        'in_progress': 'En cours',
        'completed': 'Terminé',
        'cancelled': 'Annulé'
    }
    status_fr = status_mapping.get(getattr(intervention,'status','N/A').lower(), 'N/A')

    details_data = [
        ['ID Intervention', f"#{getattr(intervention,'id','N/A')}", 'Date', date_str],
        ['Technicien', tech_name, 'Statut', status_fr],
        ['Heures', f"{hours}h", 'Déplacement', f"{travel}h"],
        ['Coût Total', f"{total_cost} FCFA", '', '']
    ]
    details_table = Table(details_data, colWidths=[1.5*inch]*4)
    details_table.setStyle(TableStyle([
        ('FONT',(0,0),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#2c3e50')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.whitesmoke),
        ('GRID',(0,0),(-1,-1),0.5,colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#f8f9fa')])
    ]))
    elements.append(details_table)
    elements.append(Spacer(1,6))

    # --- Report Section ---
    report = getattr(intervention,'report','')
    if len(report)>300: report = report[:300]+"..."
    elements.append(Paragraph("Détails de l'Intervention", section_style))
    elements.append(Paragraph(report, normal_style))
    elements.append(Spacer(1,6))

    # --- Signatures in French ---
    client_name = getattr(client,'name','N/A') if client else 'N/A'
    sig_data = [
        ['Technicien','________________','Client','________________'],
        ['Nom', tech_name, 'Nom', client_name],
        ['Date', datetime.now().strftime('%d/%m/%Y'),'Date','________________']
    ]
    sig_table = Table(sig_data, colWidths=[1.25*inch]*4)
    sig_table.setStyle(TableStyle([
        ('FONT',(0,0),(-1,-1),'Helvetica'),
        ('FONTSIZE',(0,0),(-1,-1),8),
        ('ALIGN',(0,0),(-1,-1),'LEFT'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('LINEABOVE',(0,0),(-1,0),1,colors.black)
    ]))
    elements.append(sig_table)
    elements.append(Spacer(1,6))

    # --- Footer in French ---
    footer_text = f"Rapport généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}"
    elements.append(Paragraph(footer_text, styles['Italic']))

    doc.build(elements)
    buffer.seek(0)
    return buffer