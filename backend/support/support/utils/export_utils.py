# utils/export_utils.py
import io
import base64
from django.template.loader import render_to_string
from weasyprint import HTML
from weasyprint.text.fonts import FontConfiguration
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from openpyxl import Workbook
import openpyxl
from django.utils import timezone

def export_ticket_pdf(ticket):
    """Exporte un ticket détaillé en PDF avec WeasyPrint"""
    # Préparer les données pour le template
    context = {
        'ticket': ticket,
        'now': timezone.now(),
       # 'interventions': ticket.interventions.all().order_by('intervention_date', 'start_time'),
        'images': ticket.images.all()
    }
    
    # Rendre le template HTML
    html_string = render_to_string('ticket_pdf_template.html', context)
    
    # Configurer les polices
    font_config = FontConfiguration()
    
    # Créer le PDF
    buffer = io.BytesIO()
    HTML(string=html_string).write_pdf(buffer, font_config=font_config)
    buffer.seek(0)
    
    return buffer


def export_tickets_pdf(tickets, filename="tickets_report.pdf"):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Title
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Center
    )
    title = Paragraph("Rapport des Tickets", title_style)
    elements.append(title)
    
    # Table data
    data = [['ID', 'Titre', 'Statut', 'Priorité', 'Date de création', 'Client', 'Technicien']]
    
    for ticket in tickets:
        data.append([
            str(ticket.id)[:8] + '...',
            ticket.title,
            dict(ticket.STATUS_CHOICES).get(ticket.status, ticket.status),
            dict(ticket.PRIORITY_CHOICES).get(ticket.priority, ticket.priority),
            ticket.created_at.strftime('%d/%m/%Y %H:%M'),
            f"{ticket.client.user.first_name} {ticket.client.user.last_name}",
            f"{ticket.technician.user.first_name} {ticket.technician.user.last_name}" if ticket.technician else "Non assigné"
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def export_tickets_excel(tickets, filename="tickets_report.xlsx"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    
    # Headers
    headers = ['ID', 'Titre', 'Description', 'Statut', 'Priorité', 'Date de création', 'Client', 'Technicien', 'Coût (FCFA)']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
        ws.cell(row=1, column=col_num).font = openpyxl.styles.Font(bold=True)
    
    # Data
    for row_num, ticket in enumerate(tickets, 2):
        ws.cell(row=row_num, column=1, value=str(ticket.id))
        ws.cell(row=row_num, column=2, value=ticket.title)
        ws.cell(row=row_num, column=3, value=ticket.description)
        ws.cell(row=row_num, column=4, value=dict(ticket.STATUS_CHOICES).get(ticket.status, ticket.status))
        ws.cell(row=row_num, column=5, value=dict(ticket.PRIORITY_CHOICES).get(ticket.priority, ticket.priority))
        ws.cell(row=row_num, column=6, value=ticket.created_at.strftime('%d/%m/%Y %H:%M'))
        ws.cell(row=row_num, column=7, value=f"{ticket.client.user.first_name} {ticket.client.user.last_name}")
        ws.cell(row=row_num, column=8, value=f"{ticket.technician.user.first_name} {ticket.technician.user.last_name}" if ticket.technician else "Non assigné")
       # ws.cell(row=row_num, column=9, value=f"{ticket.cost if ticket.cost else '0'} FCFA")
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
        
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer